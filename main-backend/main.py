"""
Schedulus Main Backend - API Gateway and Orchestrator

This is the main entry point for the backend API. It coordinates between
the frontend, ML Engine, and Algorithm API to provide schedule optimization.
"""

from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from contextlib import asynccontextmanager
from datetime import datetime
import uuid
import io
from openpyxl import load_workbook

from config import get_settings
from database import get_db, init_db, close_db
from models import OptimizationJob, JobStatusEnum, Lesson
from schemas import (
    OptimizationRequest,
    OptimizationJobResponse,
    TimetableResponse,
    HealthResponse,
    JobStatus,
    LessonCreate,
    LessonResponse,
)
from orchestrator import orchestrator

settings = get_settings()

# Seed data for initial lessons (used during app startup)
DEFAULT_LESSONS = [
    {"id": "l1", "subject": "Introduction to Programming", "teacher": "Dr. Smith", "student_group": "CS-1", "difficulty_weight": 0.7, "satisfaction_score": 0.85, "duration_hours": 2},
    {"id": "l2", "subject": "Data Structures", "teacher": "Dr. Johnson", "student_group": "CS-2", "difficulty_weight": 0.85, "satisfaction_score": 0.8, "duration_hours": 3},
    {"id": "l3", "subject": "Algorithms", "teacher": "Dr. Johnson", "student_group": "CS-2", "difficulty_weight": 0.9, "satisfaction_score": 0.75, "duration_hours": 3},
    {"id": "l4", "subject": "Database Systems", "teacher": "Prof. Williams", "student_group": "CS-3", "difficulty_weight": 0.75, "satisfaction_score": 0.88, "duration_hours": 2},
    {"id": "l5", "subject": "Operating Systems", "teacher": "Dr. Brown", "student_group": "CS-3", "difficulty_weight": 0.82, "satisfaction_score": 0.7, "duration_hours": 2},
    {"id": "l6", "subject": "Computer Networks", "teacher": "Dr. Davis", "student_group": "CS-3", "difficulty_weight": 0.78, "satisfaction_score": 0.82, "duration_hours": 2},
    {"id": "l7", "subject": "Linear Algebra", "teacher": "Prof. Miller", "student_group": "MATH-1", "difficulty_weight": 0.88, "satisfaction_score": 0.65, "duration_hours": 2},
    {"id": "l8", "subject": "Calculus II", "teacher": "Prof. Miller", "student_group": "MATH-1", "difficulty_weight": 0.92, "satisfaction_score": 0.6, "duration_hours": 3},
    {"id": "l9", "subject": "Statistics", "teacher": "Dr. Wilson", "student_group": "MATH-2", "difficulty_weight": 0.72, "satisfaction_score": 0.78, "duration_hours": 2},
    {"id": "l10", "subject": "Machine Learning", "teacher": "Dr. Anderson", "student_group": "CS-4", "difficulty_weight": 0.95, "satisfaction_score": 0.92, "duration_hours": 3},
]


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager for startup and shutdown events."""
    # Startup
    await init_db()
    await seed_lessons()
    yield
    # Shutdown
    await close_db()


app = FastAPI(
    title="Schedulus Main Backend",
    description="API Gateway and Orchestrator for Smart University Course Scheduling",
    version="1.0.0",
    lifespan=lifespan,
)

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/", response_model=HealthResponse)
async def root():
    """Root endpoint - health check."""
    return await health_check()


@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Health check endpoint with service status."""
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        services={
            "main_backend": "up",
            "ml_engine": settings.ml_engine_url,
            "algorithm_api": settings.algorithm_api_url,
        },
    )


@app.post("/api/schedules/optimize", response_model=OptimizationJobResponse)
async def start_optimization(
    request: OptimizationRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    """
    Start a new schedule optimization job.
    
    This endpoint initiates the optimization workflow:
    1. Enriches lessons with ML predictions
    2. Sends to Algorithm API for solving
    3. Returns job ID for status polling
    """
    job_id = uuid.uuid4()
    
    # Create job in database
    job = OptimizationJob(
        id=job_id,
        status=JobStatusEnum.PENDING,
        progress=0,
        started_at=datetime.utcnow(),
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)
    
    # Run optimization in background
    background_tasks.add_task(
        run_optimization_task,
        str(job_id),
        request,
    )
    
    return OptimizationJobResponse(
        id=str(job.id),
        status=JobStatus(job.status.value),
        progress=job.progress,
        started_at=job.started_at,
    )


async def run_optimization_task(job_id: str, request: OptimizationRequest):
    """Background task to run optimization."""
    from database import async_session_factory
    
    async with async_session_factory() as db:
        try:
            # Get job from database
            result = await db.execute(
                select(OptimizationJob).where(OptimizationJob.id == uuid.UUID(job_id))
            )
            job = result.scalar_one_or_none()
            
            if not job:
                return
            
            # Update status to running
            job.status = JobStatusEnum.RUNNING
            job.progress = 10
            await db.commit()
            
            # Convert Pydantic models to dicts
            timeslots = [ts.model_dump() for ts in request.timeslots]
            rooms = [room.model_dump() for room in request.rooms]
            lessons = [lesson.model_dump() for lesson in request.lessons]
            
            job.progress = 30
            await db.commit()
            
            # Run optimization
            optimization_result = await orchestrator.run_optimization(
                timeslots=timeslots,
                rooms=rooms,
                lessons=lessons,
            )
            
            job.progress = 90
            await db.commit()
            
            # Parse and store result
            timetable_response = parse_timetable_result(optimization_result)
            job.result = timetable_response.model_dump()
            job.status = JobStatusEnum.COMPLETED
            job.progress = 100
            job.completed_at = datetime.utcnow()
            await db.commit()
            
        except Exception as e:
            job.status = JobStatusEnum.FAILED
            job.error = str(e)
            job.completed_at = datetime.utcnow()
            await db.commit()


def parse_timetable_result(result: dict) -> TimetableResponse:
    """Parse Algorithm API result to response model."""
    from schemas import TimeslotResponse, RoomResponse, LessonResponse, ScoreResponse
    
    timeslots = [
        TimeslotResponse(
            day_of_week=ts.get("dayOfWeek", ""),
            start_time=ts.get("startTime", ""),
            end_time=ts.get("endTime", ""),
            preference_bonus=ts.get("preferenceBonus"),
        )
        for ts in result.get("timeslots", [])
    ]
    
    rooms = [
        RoomResponse(
            name=room.get("name", ""),
            capacity=room.get("capacity"),
        )
        for room in result.get("rooms", [])
    ]
    
    lessons = []
    for lesson in result.get("lessons", []):
        ts = lesson.get("timeslot")
        room = lesson.get("room")

        lessons.append(LessonResponse(
            id=lesson.get("id", ""),
            subject=lesson.get("subject", ""),
            teacher=lesson.get("teacher", ""),
            student_group=lesson.get("studentGroup", ""),
            duration_hours=lesson.get("durationHours", 2),
            difficulty_weight=lesson.get("difficultyWeight"),
            satisfaction_score=lesson.get("satisfactionScore"),
            pinned=lesson.get("pinned", False),
            timeslot=TimeslotResponse(
                day_of_week=ts.get("dayOfWeek", "") if ts else "",
                start_time=ts.get("startTime", "") if ts else "",
                end_time=ts.get("endTime", "") if ts else "",
                preference_bonus=ts.get("preferenceBonus") if ts else None,
            ) if ts else None,
            room=RoomResponse(
                name=room.get("name", "") if room else "",
                capacity=room.get("capacity") if room else None,
            ) if room else None,
        ))
    
    score = result.get("score")
    score_response = None
    if score:
        # Parse Timefold score format (e.g., "0hard/-15soft")
        if isinstance(score, str):
            hard = 0
            soft = 0
            parts = score.replace("hard", "").replace("soft", "").split("/")
            if len(parts) >= 1:
                hard = int(parts[0])
            if len(parts) >= 2:
                soft = int(parts[1])
            score_response = ScoreResponse(hard_score=hard, soft_score=soft)
        elif isinstance(score, dict):
            score_response = ScoreResponse(
                hard_score=score.get("hardScore", 0),
                soft_score=score.get("softScore", 0),
            )
    
    return TimetableResponse(
        timeslots=timeslots,
        rooms=rooms,
        lessons=lessons,
        score=score_response,
    )


@app.get("/api/schedules/jobs/{job_id}", response_model=OptimizationJobResponse)
async def get_job_status(job_id: str, db: AsyncSession = Depends(get_db)):
    """Get the status of an optimization job."""
    try:
        job_uuid = uuid.UUID(job_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid job ID format")
    
    result = await db.execute(
        select(OptimizationJob).where(OptimizationJob.id == job_uuid)
    )
    job = result.scalar_one_or_none()
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return OptimizationJobResponse(
        id=str(job.id),
        status=JobStatus(job.status.value),
        progress=job.progress,
        started_at=job.started_at,
        completed_at=job.completed_at,
        result=TimetableResponse(**job.result) if job.result else None,
        error=job.error,
    )


@app.get("/api/schedules/latest", response_model=TimetableResponse)
async def get_latest_schedule(db: AsyncSession = Depends(get_db)):
    """Get the most recently completed schedule."""
    result = await db.execute(
        select(OptimizationJob)
        .where(OptimizationJob.status == JobStatusEnum.COMPLETED)
        .where(OptimizationJob.result.isnot(None))
        .order_by(desc(OptimizationJob.completed_at))
        .limit(1)
    )
    job = result.scalar_one_or_none()
    
    if not job:
        raise HTTPException(status_code=404, detail="No completed schedules found")
    
    return TimetableResponse(**job.result)


# ========== Lessons CRUD ==========
async def seed_lessons():
    """Seed default lessons if table is empty."""
    from database import async_session_factory
    
    async with async_session_factory() as db:
        result = await db.execute(select(Lesson).limit(1))
        if result.scalar_one_or_none() is None:
            for lesson_data in DEFAULT_LESSONS:
                lesson = Lesson(
                    id=lesson_data["id"],
                    subject=lesson_data["subject"],
                    teacher=lesson_data["teacher"],
                    student_group=lesson_data["student_group"],
                    duration_hours=lesson_data.get("duration_hours", 2),
                    difficulty_weight=lesson_data["difficulty_weight"],
                    satisfaction_score=lesson_data["satisfaction_score"],
                    pinned=False,
                )
                db.add(lesson)
            await db.commit()


@app.get("/api/lessons", response_model=list[LessonResponse])
async def get_lessons(db: AsyncSession = Depends(get_db)):
    """Get all lessons."""
    result = await db.execute(select(Lesson).order_by(Lesson.id))
    lessons = result.scalars().all()
    return [LessonResponse(**lesson.to_dict()) for lesson in lessons]


@app.post("/api/lessons", response_model=LessonResponse)
async def create_lesson(lesson_data: LessonCreate, db: AsyncSession = Depends(get_db)):
    """Create a new lesson."""
    # Check if ID already exists
    existing = await db.execute(select(Lesson).where(Lesson.id == lesson_data.id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Lesson with this ID already exists")
    
    lesson = Lesson(
        id=lesson_data.id,
        subject=lesson_data.subject,
        teacher=lesson_data.teacher,
        student_group=lesson_data.student_group,
        duration_hours=lesson_data.duration_hours,
        difficulty_weight=lesson_data.difficulty_weight,
        satisfaction_score=lesson_data.satisfaction_score,
        pinned=lesson_data.pinned,
    )
    db.add(lesson)
    await db.commit()
    await db.refresh(lesson)
    return LessonResponse(**lesson.to_dict())


@app.put("/api/lessons/{lesson_id}", response_model=LessonResponse)
async def update_lesson(lesson_id: str, lesson_data: LessonCreate, db: AsyncSession = Depends(get_db)):
    """Update an existing lesson."""
    result = await db.execute(select(Lesson).where(Lesson.id == lesson_id))
    lesson = result.scalar_one_or_none()
    
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    lesson.subject = lesson_data.subject
    lesson.teacher = lesson_data.teacher
    lesson.student_group = lesson_data.student_group
    lesson.duration_hours = lesson_data.duration_hours
    lesson.difficulty_weight = lesson_data.difficulty_weight
    lesson.satisfaction_score = lesson_data.satisfaction_score
    lesson.pinned = lesson_data.pinned
    
    await db.commit()
    await db.refresh(lesson)
    return LessonResponse(**lesson.to_dict())


@app.delete("/api/lessons/{lesson_id}")
async def delete_lesson(lesson_id: str, db: AsyncSession = Depends(get_db)):
    """Delete a lesson."""
    # Remove lesson
    result = await db.execute(select(Lesson).where(Lesson.id == lesson_id))
    lesson = result.scalar_one_or_none()
    
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    await db.delete(lesson)
    await db.commit()
    
    # Also remove any scheduled occurrences of the lesson in the latest timetable
    # by marking the latest completed job result to exclude the deleted lesson.
    job_result = await db.execute(
        select(OptimizationJob)
        .where(OptimizationJob.status == JobStatusEnum.COMPLETED)
        .where(OptimizationJob.result.isnot(None))
        .order_by(desc(OptimizationJob.completed_at))
        .limit(1)
    )
    latest_job = job_result.scalar_one_or_none()
    if latest_job and latest_job.result and latest_job.result.get("lessons"):
        latest_job.result["lessons"] = [l for l in latest_job.result["lessons"] if l.get("id") != lesson_id]
        await db.commit()
    
    return {"message": "Lesson deleted"}


@app.patch("/api/lessons/{lesson_id}/pin", response_model=LessonResponse)
async def toggle_lesson_pin(lesson_id: str, db: AsyncSession = Depends(get_db)):
    """Toggle the pinned status of a lesson."""
    result = await db.execute(select(Lesson).where(Lesson.id == lesson_id))
    lesson = result.scalar_one_or_none()
    
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    lesson.pinned = not lesson.pinned
    await db.commit()
    await db.refresh(lesson)
    return LessonResponse(**lesson.to_dict())


@app.post("/api/lessons/import", response_model=list[LessonResponse])
async def import_lessons(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Import lessons from an XLSX file.
    Expected headers: id (optional),subject,teacher,student_group,duration_hours,difficulty_weight,satisfaction_score
    """
    content = await file.read()

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid XLSX file")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="XLSX file is empty")

    header_row = rows[0]
    headers = {str(col).strip().lower(): idx for idx, col in enumerate(header_row) if col}
    required_headers = {"subject", "teacher", "student_group", "duration_hours", "difficulty_weight", "satisfaction_score"}
    if not required_headers.issubset(headers.keys()):
        missing = required_headers - set(headers.keys())
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(sorted(missing))}")

    imported_lessons: list[LessonResponse] = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue

        def get_value(column: str, default=None):
            idx = headers.get(column)
            if idx is None or idx >= len(row):
                return default
            return row[idx] if row[idx] is not None else default

        def normalize_id(value):
            if value is None:
                return None
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            text = str(value).strip()
            return text or None

        def to_int(value, default: int) -> int:
            try:
                return int(value)
            except (TypeError, ValueError):
                return default

        def to_float(value, default: float) -> float:
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        lesson_id = normalize_id(get_value("id")) or f"l{uuid.uuid4().hex}"
        # Upsert: delete existing lesson with same id to keep data fresh
        existing = await db.execute(select(Lesson).where(Lesson.id == lesson_id))
        existing_lesson = existing.scalar_one_or_none()
        if existing_lesson:
            await db.delete(existing_lesson)
            await db.commit()

        lesson = Lesson(
            id=lesson_id,
            subject=str(get_value("subject", "Untitled")),
            teacher=str(get_value("teacher", "Unknown")),
            student_group=str(get_value("student_group", "")),
            duration_hours=to_int(get_value("duration_hours"), 2),
            difficulty_weight=to_float(get_value("difficulty_weight"), 0.5),
            satisfaction_score=to_float(get_value("satisfaction_score"), 0.5),
            pinned=False,
        )
        db.add(lesson)
        await db.commit()
        await db.refresh(lesson)
        imported_lessons.append(LessonResponse(**lesson.to_dict()))

    # Invalidate latest timetable since lessons changed
    latest_job_result = await db.execute(
        select(OptimizationJob)
        .where(OptimizationJob.status == JobStatusEnum.COMPLETED)
        .order_by(desc(OptimizationJob.completed_at))
        .limit(1)
    )
    latest_job = latest_job_result.scalar_one_or_none()
    if latest_job:
        latest_job.result = None
        latest_job.score = None if hasattr(latest_job, "score") else None  # safeguard
        await db.commit()

    return imported_lessons


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
